from graphical_interface import email_envio,checkbox_email,planilha_pre
from functions import pass_random, pin_random, meeting_random
import csv
import openpyxl

print(planilha_pre,email_envio,checkbox_email)

name=lastname=number=email=data=''

f = open('extensions.csv','w', encoding='UTF8',newline='')
cabecalho = ['Number','FirstName','LastName','EmailAddress','MobileNumber','AuthID','AuthPassword','WebMeetingFriendlyName','WebMeetingPrivateRoom','ClickToCall','ClickToCallFriendlyName','WebMeetingAcceptReject','EnableVoicemail','VMNoPin','VMPlayCallerID','PIN','VMPlayMsgDateTime','VMEmailOptions','QueueStatus','OutboundCallerID','SIPID','DeliverAudio','SupportReinvite','SupportReplaces','EnableSRTP','ManagementAccess','ReporterAccess','WallboardAccess','TurnOffMyPhone','HideFWrules','CanSeeRecordings','CanDeleteRecordings','RecordCalls','CallScreening','EmailMissedCalls','Disabled','DisableExternalCalls','AllowLanOnly','BlockRemoteTunnel','PinProtect','MAC_0','InterfaceIP_0','UseTunnel','DND','UseCTI','StartupScreen','HotelModuleAccess','DontShowExtInPHBK','DeskphoneWebPass','SrvcAccessPwd','VoipAdmin','SysAdmin','SecureSIP','PhoneModel14','PhoneTemplate14','CustomTemplate','PhoneSettings','AllowAllRecordings','PushExtension','Integration','AllowOwnRecordings','RecordExternalCallsOnly','DID','SMS','PhoneSysAdmin']
writer = csv.writer(f)
writer.writerow(cabecalho)  


path = planilha_pre
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row=sheet_obj.max_row
max_col = sheet_obj.max_column

for line in range (6, m_row+1):
    for i in range(7, 11):
        cell_obj = sheet_obj.cell(row = line, column = i)         
        if i ==7:
            number = cell_obj.value          
        elif i==8:
            name=cell_obj.value
        elif i==9:
            lastname=cell_obj.value
        elif i==10:            
            email=cell_obj.value
            if number!=None:
                senha = pass_random()
                meeting_address= meeting_random()
                pin_number=pin_random()
                if checkbox_email == True:
                    email = email_envio
                data=[number,name,lastname,email,'',senha,senha,meeting_address,0,0,'',0,0,1,0,pin_number,0,0,1,'','',0,1,1,0,0,0,'',0,0,0,0,0,0,0,0,0,0,0,0,'','',1,0,0,0,'',0,senha,senha,0,0,0,'','','','',0,1,'',1,0,'','',0]   
                with open('extensions.csv', 'a', encoding='UTF8',newline='') as f:
                    writer = csv.writer(f)                                                                                          
                    writer.writerow(data)
                print(data)
                        

    

